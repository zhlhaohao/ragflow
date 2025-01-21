#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

from openpyxl import load_workbook
import sys
from io import BytesIO
from rag.nlp import find_codec
from api.utils import ic


class RAGFlowExcelParser:
    def html(self, fnm, chunk_rows=256):
        """
        将Excel文件转换为HTML格式的表格。

        参数:
        fnm (str or bytes): Excel文件的路径或二进制数据。
        chunk_rows (int): 每个表格块包含的行数，默认为256。

        返回:
        list: 包含HTML表格字符串的列表。
        """
        # 如果fnm是字符串，则认为是文件路径，加载Excel文件
        if isinstance(fnm, str):
            wb = load_workbook(fnm, data_only=True)   #F8080
        else:
            # 否则认为是二进制数据，使用BytesIO将其转换为文件对象再加载Excel文件
            wb = load_workbook(BytesIO(fnm), data_only=True)  # F8080

        tb_chunks = []
        # 遍历Excel文件中的每个工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            # F8080 检测表头行
            header_row_index = self.detect_header_row(rows)
            # 构建表头行的HTML字符串
            tb_rows_0 = "<tr>"
            for t in list(rows[header_row_index]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"
            # 将数据分块，每块包含chunk_rows行（不包括表头）
            for chunk_i in range((len(rows) - header_row_index - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
                    rows[header_row_index + 1 + chunk_i * chunk_rows : header_row_index + 1 + (chunk_i + 1) * chunk_rows]
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)
        return tb_chunks

    def __call__(self, fnm):
        """
        解析Excel文件并返回格式化的字符串列表。

        参数:
        fnm (str or bytes): Excel文件的路径或二进制数据。

        返回:
        list: 包含格式化字符串的列表。
        """
        # 如果fnm是字符串，则认为是文件路径，加载Excel文件
        if isinstance(fnm, str):
            wb = load_workbook(fnm, data_only=True)  #F8080 不读公式
        else:
            # 否则认为是二进制数据，使用BytesIO将其转换为文件对象再加载Excel文件
            wb = load_workbook(BytesIO(fnm), data_only=True) #F8080 不读公式
        res = []
        # 遍历Excel文件中的每个工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            # F8080 检测表头行
            header_row_index = self.detect_header_row(rows)
            ti = list(rows[header_row_index])
            # 遍历数据行（不包括表头）
            for r in list(rows[header_row_index + 1:]):
                # 用于存储当前行的字段
                fields = []
                # 遍历当前行的单元格，c代表单元格,i代表单元格序号
                for i, c in enumerate(r):
                    # ic(c.value)
                    cell_value = c.value
                    if not cell_value:
                        # F8080 检查是否为合并单元格
                        for merged_range in ws.merged_cells.ranges:
                            if c.coordinate in merged_range:
                                # 获取合并单元格的值
                                cell_value = ws[merged_range.start_cell.coordinate].value
                                break
                    if not cell_value:
                        continue
                    # 获取表头单元格的值
                    t = str(ti[i].value) if i < len(ti) else ""
                    # 添加当前单元格的值
                    t += ("：" if t else "") + str(cell_value)
                    # ic(t)
                    # 将字段添加到列表中
                    fields.append(t)
                  # 将字段用分号连接成字符串
                line = "; ".join(fields)
                # 如果工作表名不包含"sheet"，则添加工作表名
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                # ic(line)
                # 将结果字符串添加到列表中
                res.append(line)
        # ic(res)
        return res

    @staticmethod
    def row_number(fnm, binary):
        """
        计算Excel文件中的总行数。

        参数:
        fnm (str): 文件名。
        binary (bytes): 文件的二进制数据。

        返回:
        int: 文件中的总行数。
        """
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = load_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))

    def detect_header_row(self, rows):
        """
        F8080 检测表头在哪一行。如果 N 行非空单元格数量小于 N+1 行非空单元格数量，则 N+1 为表头行。

        参数:
        rows (list): Excel工作表的所有行。

        返回:
        int: 表头行的索引。
        """
        for i in range(len(rows) - 1):
            row1 = [cell.value for cell in rows[i] if cell.value is not None]
            row2 = [cell.value for cell in rows[i + 1] if cell.value is not None]
            if len(row1) < len(row2) and i < 3:
                return i+1
        return 0

if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
