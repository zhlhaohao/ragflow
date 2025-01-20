import unittest
from deepdoc.parser.excel_parser import RAGFlowExcelParser
import os
from openpyxl import load_workbook
from icecream import ic

class TestRAGFlowExcelParser(unittest.TestCase):
    """Ctrl+Shift+Y 打开 Debug Console,右上角选择Debug Unit Test,然后点击函数左边的调试按钮就可以看到运行结果了

    Args:
        unittest (_type_): _description_
    """
    def setUp(self):
        self.parser = RAGFlowExcelParser()

    def test_detect_header_row(self):
        """测试 detect_header_row 方法，验证表头行索引是否正确"""
        # 获取测试文件的绝对路径
        test_file_path = os.path.join(os.path.dirname(__file__), 'test_data', '广东网络研发中心通信录202501.xlsx')

        wb = load_workbook(test_file_path)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)

            # 检测表头行
            header_row_index = self.parser.detect_header_row(rows)

            # 断言表头行索引为0
            self.assertEqual(header_row_index, 1)
            break

    def test_call_method(self):
        """测试 __call__ 方法，验证解析结果是否正确"""
        # 获取测试文件的绝对路径
        test_file_path = os.path.join(os.path.dirname(__file__), 'test_data', '广东网络研发中心通信录202501.xlsx')

        # 调用 __call__ 方法
        result = self.parser(test_file_path)

        # 打印 result
        ic(result)

if __name__ == '__main__':
    unittest.main()